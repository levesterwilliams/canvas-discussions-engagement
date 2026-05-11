# course_config_loader.py
# Author: Levester Williams
# 1 May 2026

"""Course configuration loading helpers.

This module reads the workbook that drives the reporting pipeline and
normalizes each usable row into a tuple containing the course name,
Canvas course ID, and destination Box folder ID.
"""

from datetime import date, datetime
from pathlib import Path
from typing import List, Set, Tuple

from openpyxl import load_workbook


def _get_course_config_sheet(workbook):
    """Return the worksheet that contains course configuration rows.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        Workbook containing the reporting configuration sheets.

    Returns
    -------
    openpyxl worksheet
        Worksheet used for course row extraction.

    Notes
    -----
    A sheet named ``courses`` is preferred. If that sheet is absent,
    the active worksheet is used for backward compatibility.
    """
    return workbook["courses"] if "courses" in workbook.sheetnames else workbook.active


def read_course_config_xlsx(path: str | Path) -> List[Tuple[str, str, str]]:
    """Read course configuration rows from an Excel workbook.

    Parameters
    ----------
    path : str or pathlib.Path
        Path to the course configuration workbook.

    Returns
    -------
    list of tuple of str
        Normalized rows in the form
        ``(course_name, course_id, box_folder_id)``.

    Notes
    -----
    Rows missing a course ID or Box folder ID are skipped.
    """
    workbook = load_workbook(filename=Path(path))
    worksheet = _get_course_config_sheet(workbook)

    rows: List[Tuple[str, str, str]] = []
    for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        course_name, course_id, box_folder_id = row

        if not course_id or not box_folder_id:
            print(f"Skipping row {index}: missing required fields")
            continue

        rows.append((
            str(course_name or "Unknown Course"),
            str(course_id),
            str(box_folder_id),
        ))

    return rows


def read_scheduled_run_dates(path: str | Path) -> Set[date]:
    """Read scheduled report run dates from a schedule workbook.

    Parameters
    ----------
    path : str or pathlib.Path
        Path to the schedule workbook.

    Returns
    -------
    set of datetime.date
        Set of allowed run dates loaded from the workbook.

    Notes
    -----
    A sheet named ``schedule`` is preferred. If it is absent, the active
    worksheet is used so a simple single-sheet workbook also works.
    """
    workbook = load_workbook(filename=Path(path), data_only=True)
    worksheet = workbook["schedule"] if "schedule" in workbook.sheetnames else workbook.active
    run_dates: Set[date] = set()

    for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_value = row[0] if row else None
        if raw_value in (None, ""):
            continue

        if isinstance(raw_value, datetime):
            run_dates.add(raw_value.date())
            continue

        if isinstance(raw_value, date):
            run_dates.add(raw_value)
            continue

        if isinstance(raw_value, str):
            try:
                run_dates.add(date.fromisoformat(raw_value.strip()))
            except ValueError:
                print(f"Skipping schedule row {index}: invalid ISO date '{raw_value}'")
            continue

        print(f"Skipping schedule row {index}: unsupported date value '{raw_value}'")

    return run_dates

