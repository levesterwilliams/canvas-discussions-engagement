# canvas_discussions_engagement.py
# Author: Levester Williams
# 31 July 2024

"""Canvas discussion analytics client and Excel export helpers."""

from __future__ import annotations

import json
import os
import re
import tempfile
import time
from collections import OrderedDict
from json import JSONDecodeError
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class CanvasDiscussions:
    """Collect discussion participation data from Canvas LMS.

    Parameters
    ----------
    server_type : str
        Canvas environment key, such as ``"LPS_Production"``.
    enrollment : str
        Enrollment role key, such as ``"Student"``, ``"TA"``, or
        ``"Teacher"``.
    course_number : str
        Canvas course identifier.

    Attributes
    ----------
    discussions_meta : list of dict
        Ordered metadata for published discussion topics.
    module_discussion_map : dict of str to list of int
        Mapping of module names to discussion topic IDs.
    course_name : str
        Human-readable course name used in the generated workbook name.
    """

    server_url = {
        "LPS_Production": "https://canvas.upenn.edu/",
        "LPS_Test": "https://upenn.test.instructure.com/",
    }

    enrollment_type = {
        "Student": "StudentEnrollment",
        "TA": "TaEnrollment",
        "Teacher": "TeacherEnrollment",
    }

    course_name = "Unknown"

    def __init__(self, server_type: str, enrollment: str, course_number: str) -> None:
        """Initialize a Canvas discussion client."""
        self.server_type = server_type
        self.enrollment = enrollment
        self.course_num = course_number
        self.discussions_meta: List[Dict[str, object]] = []
        self.module_discussion_map: Dict[str, List[int]] = {}

    def _get_token(self) -> dict:
        """Read Canvas API credentials from the environment.

        Returns
        -------
        dict
            Parsed ``CANVAS_API_CRED`` payload keyed by server name.

        Raises
        ------
        RuntimeError
            Raised when the environment variable is missing or invalid.
        """
        try:
            raw_cred = os.getenv("CANVAS_API_CRED")
            if raw_cred is None:
                raise RuntimeError("Environment variable CANVAS_API_CRED does not exist.")
            return json.loads(raw_cred)
        except json.JSONDecodeError as exc:
            raise RuntimeError("CANVAS_API_CRED contains invalid JSON.") from exc

    def _headers(self) -> dict:
        """Construct Canvas API request headers.

        Returns
        -------
        dict
            Headers containing JSON content type and bearer token.
        """
        token = self._get_token()
        return {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {token[self.server_type]}",
        }

    def _get_server_url(self) -> str:
        """Return the base URL for the configured Canvas environment.

        Returns
        -------
        str
            Canvas server base URL.
        """
        return self.server_url[self.server_type]

    def get_enrollment_type(self) -> str:
        """Return the Canvas enrollment type string for the current role.

        Returns
        -------
        str
            Canvas enrollment type value used in API queries.
        """
        return self.enrollment_type[self.enrollment]

    def set_enrollment_type(self, enrollment_type: str) -> str:
        """Update the enrollment role used for downstream API calls.

        Parameters
        ----------
        enrollment_type : str
            Role key to apply, such as ``"Teacher"`` or ``"TA"``.

        Returns
        -------
        str
            Canvas enrollment type string for the new role.
        """
        self.enrollment = enrollment_type
        return self.enrollment_type[self.enrollment]

    def _get_next_page_url(self, link_header: Optional[str]) -> str:
        """Extract the next-page URL from a Canvas pagination header.

        Parameters
        ----------
        link_header : str or None
            Raw HTTP ``Link`` header value.

        Returns
        -------
        str
            Next page URL, or an empty string when pagination ends.
        """
        if link_header:
            links = link_header.split(",")
            for link in links:
                if 'rel="next"' in link:
                    return link.split(";")[0].strip("<> ")
        return ""

    def _paged_get(self, url: str) -> List[dict]:
        """Perform a paginated Canvas GET request.

        Parameters
        ----------
        url : str
            Initial Canvas API endpoint URL.

        Returns
        -------
        list of dict
            Aggregated JSON items across all returned pages.

        Notes
        -----
        HTTP 500 responses are retried up to three times per page.
        """
        items: List[dict] = []
        page_url = url
        max_retries = 3
        retry_delay = 2

        while page_url:
            for _ in range(max_retries):
                response = requests.get(page_url, headers=self._headers(), timeout=60)
                if response.status_code == 200:
                    try:
                        chunk = response.json()
                    except JSONDecodeError:
                        print("Failed to decode JSON data from Canvas response.")
                        return []

                    if isinstance(chunk, list):
                        items.extend(chunk)
                    elif isinstance(chunk, dict):
                        items.append(chunk)
                    else:
                        print("Unexpected Canvas API response format.")
                        return []

                    page_url = self._get_next_page_url(response.headers.get("Link"))
                    break

                if response.status_code == 500:
                    time.sleep(retry_delay)
                    continue

                print(f"Unexpected Canvas error ({response.status_code}): {response.text}")
                return []
            else:
                print("Max retries reached. Could not complete paged GET.")
                return []

        return items

    def get_enrollees(self, course_id: str) -> List[Tuple[int, str]]:
        """Retrieve course enrollees for the active enrollment role.

        Parameters
        ----------
        course_id : str
            Canvas course identifier.

        Returns
        -------
        list of tuple
            ``(user_id, sortable_name)`` pairs for matching enrollments.
        """
        enrollments_url = (
            f"{self._get_server_url()}api/v1/courses/{course_id}/enrollments"
            f"?type[]={self.get_enrollment_type()}&per_page=100"
        )
        items = self._paged_get(enrollments_url)
        filtered = [
            enrollment for enrollment in items
            if isinstance(enrollment, dict) and enrollment.get("type") == self.get_enrollment_type()
        ]

        enrollments: List[Tuple[int, str]] = []
        for enrollment in filtered:
            user = enrollment.get("user", {})
            user_id = user.get("id", "Unknown")
            name = user.get("sortable_name", "Unknown")
            if isinstance(user_id, int) and isinstance(name, str):
                enrollments.append((user_id, name.strip()))
        return enrollments

    def _get_full_topic_view(self, course_id: str, topic_id: int) -> dict:
        """Retrieve the full Canvas discussion topic view.

        Parameters
        ----------
        course_id : str
            Canvas course identifier.
        topic_id : int
            Discussion topic identifier.

        Returns
        -------
        dict
            Topic payload including participants when accessible,
            otherwise an empty dictionary.
        """
        url = f"{self._get_server_url()}api/v1/courses/{course_id}/discussion_topics/{topic_id}/view"
        response = requests.get(url, headers=self._headers(), timeout=60)
        if response.status_code == 200:
            try:
                return response.json()
            except JSONDecodeError:
                print("Failed to decode JSON from topic view response.")
                return {}
        if response.status_code == 403:
            return {}

        print(f"Error fetching full topic view: {response.status_code}, {response.text}")
        return {}

    def _process_full_topic_view(
        self,
        course_id: str,
        topic_id: int,
        enrollee_discussion_data: Dict[str, List[bool]],
        topic_title: str,
        enrollees_in_course: List[Tuple[int, str]],
        list_topic_titles: List[str],
    ) -> None:
        """Mark participant activity for a single discussion topic.

        Parameters
        ----------
        course_id : str
            Canvas course identifier.
        topic_id : int
            Discussion topic identifier.
        enrollee_discussion_data : dict of str to list of bool
            Mutable participation matrix keyed by enrollee name.
        topic_title : str
            Discussion title used to locate the column index.
        enrollees_in_course : list of tuple
            Known course enrollees as ``(user_id, sortable_name)`` pairs.
        list_topic_titles : list of str
            Ordered list of discussion titles matching the matrix columns.
        """
        topic_view = self._get_full_topic_view(course_id, topic_id)
        if not topic_view:
            return

        id_to_name = {user_id: name for user_id, name in enrollees_in_course}
        try:
            discussion_index = list_topic_titles.index(topic_title)
        except ValueError:
            return

        for participant in topic_view.get("participants", []):
            participant_id = participant.get("id")
            if participant_id in id_to_name:
                enrollee_name = id_to_name[participant_id]
                enrollee_discussion_data[enrollee_name][discussion_index] = True

    def get_course_discussion_data(
        self,
        course_id: str,
        enrollees_in_course: List[Tuple[int, str]],
    ) -> Tuple[OrderedDict[str, List[bool]], List[str]]:
        """Build the course discussion participation matrix.

        Parameters
        ----------
        course_id : str
            Canvas course identifier.
        enrollees_in_course : list of tuple
            Course enrollee records as ``(user_id, sortable_name)`` pairs.

        Returns
        -------
        collections.OrderedDict of str to list of bool
            Participation matrix keyed by sortable enrollee name.
        list of str
            Ordered discussion titles corresponding to each matrix column.

        Notes
        -----
        Published discussions are sorted by their effective posted date
        before the matrix is assembled.
        """
        page_url = f"{self._get_server_url()}api/v1/courses/{course_id}/discussion_topics?per_page=100"
        discussions: List[Tuple[str, int, str]] = []

        while page_url:
            response = requests.get(page_url, headers=self._headers(), timeout=60)
            if response.status_code != 200:
                print(f"Unexpected Canvas error ({response.status_code}): {response.text}")
                break

            try:
                discussion_topics = response.json()
            except json.JSONDecodeError:
                print("Failed to decode JSON data from discussion topics response.")
                return OrderedDict(), []

            for topic in discussion_topics:
                if topic.get("published", False):
                    topic_title = topic.get("title", "Unknown Title")
                    topic_id = topic.get("id")
                    topic_posted_date = (
                        topic.get("last_reply_at")
                        or topic.get("posted_at")
                        or topic.get("created_at")
                        or "1900-01-01T00:00:00Z"
                    )
                    if isinstance(topic_id, int):
                        discussions.append((topic_posted_date, topic_id, topic_title))

            page_url = self._get_next_page_url(response.headers.get("Link"))

        discussions.sort(key=lambda item: item[0] or "1900-01-01T00:00:00Z")
        self.discussions_meta = [{"id": topic_id, "title": title} for _, topic_id, title in discussions]
        topic_titles = [str(meta["title"]) for meta in self.discussions_meta]

        enrollee_discussion_data: Dict[str, List[bool]] = {
            enrollee_name: [False] * len(self.discussions_meta)
            for _, enrollee_name in enrollees_in_course
        }

        for _, topic_id, topic_title in discussions:
            self._process_full_topic_view(
                course_id=course_id,
                topic_id=topic_id,
                enrollee_discussion_data=enrollee_discussion_data,
                topic_title=topic_title,
                enrollees_in_course=enrollees_in_course,
                list_topic_titles=topic_titles,
            )

        self._build_module_discussion_map(course_id)
        return OrderedDict(sorted(enrollee_discussion_data.items())), topic_titles

    def _get_modules(self, course_id: str) -> List[dict]:
        """Retrieve modules for a course.

        Parameters
        ----------
        course_id : str
            Canvas course identifier.

        Returns
        -------
        list of dict
            Module payloads returned by Canvas.
        """
        url = f"{self._get_server_url()}api/v1/courses/{course_id}/modules?per_page=100"
        return self._paged_get(url)

    def _get_module_items(self, course_id: str, module_id: int) -> List[dict]:
        """Retrieve items for a specific Canvas module.

        Parameters
        ----------
        course_id : str
            Canvas course identifier.
        module_id : int
            Canvas module identifier.

        Returns
        -------
        list of dict
            Module item payloads returned by Canvas.
        """
        url = f"{self._get_server_url()}api/v1/courses/{course_id}/modules/{module_id}/items?per_page=100"
        return self._paged_get(url)

    def _build_module_discussion_map(self, course_id: str) -> None:
        """Map Canvas module names to discussion topic IDs.

        Parameters
        ----------
        course_id : str
            Canvas course identifier.

        Notes
        -----
        Only module items of type ``"Discussion"`` are included.
        """
        self.module_discussion_map = {}
        modules = self._get_modules(course_id)
        if not modules:
            print("No Canvas modules returned. Per-module sheets will be skipped.")
            return

        for module in modules:
            module_name = (module.get("name") or f"Module {module.get('id', 'Unknown')}").strip()
            module_id = module.get("id")
            if not isinstance(module_id, int):
                continue

            items = self._get_module_items(course_id, module_id)
            for item in items:
                if item.get("type") == "Discussion":
                    self.module_discussion_map.setdefault(module_name, []).append(item["content_id"])

    def _output_basepath(self) -> Path:
        """Resolve the directory used for workbook output.

        Returns
        -------
        pathlib.Path
            Existing directory that can be used for generated reports.

        Notes
        -----
        ``OUTPUT_DIRECTORY`` is used when defined; otherwise a temporary
        directory under the host runtime is created.
        """
        output_directory = os.getenv("OUTPUT_DIRECTORY")
        if output_directory:
            base_path = Path(output_directory)
        else:
            base_path = Path(tempfile.gettempdir()) / "canvas-discussion-reports"

        base_path.mkdir(parents=True, exist_ok=True)
        return base_path

    def _role_label(self) -> str:
        """Return the report label for the current enrollment role.

        Returns
        -------
        str
            Filename-friendly role label.
        """
        return "students" if self.get_enrollment_type() == "StudentEnrollment" else "instructors"

    def _unique_sheet_name(self, workbook: Workbook, base: str) -> str:
        """Create a workbook-safe and unique worksheet name.

        Parameters
        ----------
        workbook : openpyxl.Workbook
            Workbook that will receive the sheet.
        base : str
            Preferred sheet title.

        Returns
        -------
        str
            Unique worksheet name that satisfies Excel restrictions.
        """
        invalid = set("[]:*?/\\")
        cleaned = "".join(character for character in base if character not in invalid).strip().rstrip(".")
        if not cleaned:
            cleaned = "Sheet"

        name = cleaned[:31]
        if name not in workbook.sheetnames:
            return name

        index = 2
        while True:
            candidate = f"{name[:31 - len(str(index)) - 1]} {index}"
            if candidate not in workbook.sheetnames:
                return candidate
            index += 1

    def _autowidth(self, worksheet, max_rows: int, max_cols: int) -> None:
        """Apply a simple content-based width heuristic to worksheet columns.

        Parameters
        ----------
        worksheet : openpyxl worksheet
            Worksheet whose columns should be resized.
        max_rows : int
            Maximum populated row to inspect.
        max_cols : int
            Maximum populated column to inspect.
        """
        for column_number in range(1, max_cols + 1):
            column_letter = get_column_letter(column_number)
            max_length = 10
            for row_number in range(1, max_rows + 1):
                value = worksheet.cell(row=row_number, column=column_number).value
                max_length = max(max_length, len(str(value)) if value is not None else 0)
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

    def _safe_course_name(self) -> str:
        """Return a filesystem-safe version of the course name.

        Returns
        -------
        str
            Sanitized course name suitable for output filenames.
        """
        return re.sub(r'[<>:"/\\|?*]+', "_", self.course_name).strip() or "Unknown Course"

    def write_module_breakdown_xlsx(self, enrollee_discussion_data: Dict[str, List[bool]]) -> Path:
        """Write discussion participation data to an Excel workbook.

        Parameters
        ----------
        enrollee_discussion_data : dict of str to list of bool
            Participation matrix keyed by enrollee name.

        Returns
        -------
        pathlib.Path
            Path to the saved workbook.

        Raises
        ------
        RuntimeError
            Raised when no discussion metadata is available to write.

        Notes
        -----
        The workbook includes one ``Overview_All`` sheet and additional
        sheets for any module containing discussion items.
        """
        if not self.module_discussion_map:
            self._build_module_discussion_map(self.course_num)

        if not self.discussions_meta:
            raise RuntimeError("No discussion data to write.")

        id_to_index: Dict[int, int] = {
            int(meta["id"]): index
            for index, meta in enumerate(self.discussions_meta)
            if isinstance(meta.get("id"), int)
        }

        workbook = Workbook()
        workbook.remove(workbook.active)

        overview = workbook.create_sheet(title="Overview_All")
        overview_headers = ["Name"] + [str(meta["title"]) for meta in self.discussions_meta]
        for column_number, header in enumerate(overview_headers, start=1):
            overview.cell(row=1, column=column_number, value=header)

        row_number = 2
        for enrollee_name, participations in enrollee_discussion_data.items():
            overview.cell(row=row_number, column=1, value=enrollee_name)
            for column_number, participated in enumerate(participations, start=2):
                overview.cell(row=row_number, column=column_number, value="Yes" if participated else "No")
            row_number += 1
        self._autowidth(overview, row_number - 1, len(overview_headers))

        for module_name in sorted(self.module_discussion_map.keys()):
            topic_ids = self.module_discussion_map.get(module_name, [])
            column_indexes = [id_to_index[topic_id] for topic_id in topic_ids if topic_id in id_to_index]
            if not column_indexes:
                continue

            worksheet = workbook.create_sheet(title=self._unique_sheet_name(workbook, module_name))
            headers = ["Name"] + [str(self.discussions_meta[index]["title"]) for index in column_indexes]
            for column_number, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=column_number, value=header)

            row_number = 2
            for enrollee_name, participations in enrollee_discussion_data.items():
                worksheet.cell(row=row_number, column=1, value=enrollee_name)
                for column_number, overall_index in enumerate(column_indexes, start=2):
                    worksheet.cell(
                        row=row_number,
                        column=column_number,
                        value="Yes" if participations[overall_index] else "No",
                    )
                row_number += 1
            self._autowidth(worksheet, row_number - 1, len(headers))

        xlsx_path = self._output_basepath() / f"{self._safe_course_name()}__{self._role_label()}_discussions.xlsx"
        workbook.save(xlsx_path)
        print(f"XLSX written: {xlsx_path}")
        return xlsx_path
