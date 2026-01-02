# canvas_discussions_engagement.py
# Author: Levester Williams
# 31 July 2024
#



from __future__ import annotations

import requests
import json
import sys
import os
import time
from typing import Dict, List, Tuple, Optional, OrderedDict as OrderedDictType
from json import JSONDecodeError
from pathlib import Path
from json_freader import JSONfreader
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class CanvasDiscussions:
    server_url = {
        'LPS_Production': 'https://canvas.upenn.edu/',
        'LPS_Test': 'https://upenn.test.instructure.com/',
    }

    enrollment_type = {
        'Student': 'StudentEnrollment',
        'TA': 'TaEnrollment',
        'Teacher': 'TeacherEnrollment',
    }

    course_name = "Unknown"

    def __init__(self, server_type: str, enrollment: str, course_number: str) -> None:
        self.server_type = server_type
        self.enrollment = enrollment
        self.course_num = course_number
        self.discussions_meta: List[Dict[str, object]] = []          # [{"id": int, "title": str}]
        self.module_discussion_map: Dict[str, List[int]] = {}         # module_name -> [discussion_topic_id, ...]

    # ---------- Credentials & headers
    def get_token(self=None) -> dict:
        return self.get_cred_env_var()

    def get_cred_json(self=None) -> dict:
        reader = JSONfreader()
        json_file_path = ""
        try:
            cred = reader.load_json_file(json_file_path)
        except FileNotFoundError:
            print("The credentials file cred.json was not found")
            sys.exit(1)
        except RuntimeError:
            print("The credentials file cred.json contains invalid JSON.")
            sys.exit(1)
        except Exception as e:
            print(f"Unexpected error: {e}")
            sys.exit(1)
        return cred

    def get_cred_env_var(self=None) -> dict:
        try:
            cred = json.loads(os.getenv('CANVAS_API_CRED'))
        except KeyError:
            print("Environment variable CANVAS_API_CRED does not exist.")
            sys.exit(1)
        except json.JSONDecodeError:
            print("Contains invalid JSON.")
            sys.exit(1)
        except TypeError:
            print("Invalid type: expected a JSON string.")
            sys.exit(1)
        except Exception as e:
            print(f"Unexpected error: {e}")
            sys.exit(1)
        return cred

    def headers(self) -> dict:
        token = self.get_token()
        return {
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {token[self.server_type]}',
        }

    # ---------- Simple getters/setters
    def get_server_url(self=None) -> str:
        return self.server_url[self.server_type]

    def get_enrollment_type(self=None) -> str:
        return self.enrollment_type[self.enrollment]

    def set_enrollment_type(self, enrollment_type: str) -> str:
        self.enrollment = enrollment_type
        return self.enrollment_type[self.enrollment]

    # ---------- Canvas API helpers
    def get_next_page_url(self, link_header: Optional[str]) -> str:
        if link_header:
            links = link_header.split(',')
            for link in links:
                if 'rel="next"' in link:
                    return link.split(';')[0].strip('<> ')
        return ""

    def _paged_get(self, url: str) -> List[dict]:
        items: List[dict] = []
        page_url = url
        max_retries = 3
        retry_delay = 2
        while page_url:
            for _ in range(max_retries):
                resp = requests.get(page_url, headers=self.headers())
                if resp.status_code == 200:
                    try:
                        chunk = resp.json()
                        if isinstance(chunk, list):
                            items.extend(chunk)
                        elif isinstance(chunk, dict):
                            items.append(chunk)
                        else:
                            print("Error: Unexpected API response format")
                            return []
                        page_url = self.get_next_page_url(resp.headers.get('Link'))
                        break
                    except JSONDecodeError:
                        print("Failed to decode JSON data from response")
                        return []
                elif resp.status_code == 500:
                    time.sleep(retry_delay)
                else:
                    print(f"Unexpected error ({resp.status_code}): {resp.text}")
                    return []
            else:
                print("Max retries reached. Could not complete paged GET.")
                return []
        return items

    # ---------- Enrollments
    def get_enrollees(self, course_id: str) -> List[Tuple[int, str]]:
        enrollments_url = (
            f'{self.get_server_url()}api/v1/courses/{course_id}/enrollments'
            f'?type[]={self.get_enrollment_type()}&per_page=100'
        )
        items = self._paged_get(enrollments_url)
        filtered = [
            e for e in items
            if isinstance(e, dict) and e.get('type') == self.get_enrollment_type()
        ]
        enrollments: List[Tuple[int, str]] = []
        for e in filtered:
            uid = e.get('user', {}).get('id', 'Unknown')
            name = e.get('user', {}).get('sortable_name', 'Unknown')
            if isinstance(uid, int) and isinstance(name, str):
                enrollments.append((uid, name.strip()))
        return enrollments

    # ---------- Discussions
    def get_full_topic_view(self, course_id: str, topic_id: int) -> dict:
        # Fix accidental double slash
        url = (
            f'{self.get_server_url()}api/v1/courses/{course_id}'
            f'/discussion_topics/{topic_id}/view'
        )
        response = requests.get(url, headers=self.headers())
        if response.status_code == 200:
            try:
                return response.json()
            except JSONDecodeError:
                print("Failed to decode JSON from response")
                return {}
        elif response.status_code == 403:
            return {}
        else:
            print(f"Error fetching full topic view: {response.status_code}, {response.text}")
            return {}

    def process_full_topic_view(
        self,
        course_id: str,
        topic_id: int,
        enrollee_discussion_data: Dict[str, List[bool]],
        topic_title: str,
        enrollees_in_course: List[Tuple[int, str]],
        list_topic_titles: List[str],
    ) -> None:
        topic_view = self.get_full_topic_view(course_id, topic_id)
        if not topic_view:
            return
        id_to_name = {uid: name for uid, name in enrollees_in_course}
        try:
            idx = list_topic_titles.index(topic_title)
        except ValueError:
            return
        for participant in topic_view.get('participants', []):
            pid = participant.get('id')
            if pid in id_to_name:
                enrollee_name = id_to_name[pid]
                enrollee_discussion_data[enrollee_name][idx] = True

    def get_course_discussion_data(
        self,
        course_id: str,
        enrollees_in_course: List[Tuple[int, str]],
    ) -> Tuple[OrderedDictType[str, List[bool]], List[str]]:
        page_url = f'{self.get_server_url()}api/v1/courses/{course_id}/discussion_topics?per_page=100'
        discussions: List[Tuple[str, int, str]] = []
        while page_url:
            response = requests.get(page_url, headers=self.headers())
            if response.status_code == 200:
                try:
                    discussion_topics = response.json()
                    for topic in discussion_topics:
                        if topic.get('published', False):
                            topic_title = topic.get('title', 'Unknown Title')
                            topic_id = topic.get('id', None)
                            topic_posted_date = (
                                topic.get('last_reply_at')
                                or topic.get('posted_at')
                                or topic.get('created_at')
                                or '1900-01-01T00:00:00Z'
                            )
                            if isinstance(topic_id, int):
                                discussions.append((topic_posted_date, topic_id, topic_title))
                    page_url = self.get_next_page_url(response.headers.get('Link'))
                except json.JSONDecodeError:
                    print("Failed to decode JSON data from response")
                    return OrderedDict(), []
                except KeyError:
                    print("Key error in processing discussion topics")
                    return OrderedDict(), []
            else:
                print(f"Unexpected error ({response.status_code}): {response.text}")
                page_url = None

        discussions.sort(key=lambda x: x[0] or '1900-01-01T00:00:00Z')
        self.discussions_meta = [{"id": tid, "title": title} for _, tid, title in discussions]
        list_topic_titles = [m["title"] for m in self.discussions_meta]

        enrollee_discussion_data: Dict[str, List[bool]] = {
            enrollee_name: [False] * len(self.discussions_meta)
            for _, enrollee_name in enrollees_in_course
        }

        for _, topic_id, topic_title in discussions:
            self.process_full_topic_view(
                course_id,
                topic_id,
                enrollee_discussion_data,
                topic_title,
                enrollees_in_course,
                list_topic_titles,
            )

        # Build the module map here (normal path)
        self.build_module_discussion_map(course_id)
        print(f"Map keys for module are "
              f"{self.module_discussion_map.keys()}")
        ordered_by_sortable_name = OrderedDict(sorted(enrollee_discussion_data.items()))
        return ordered_by_sortable_name, list_topic_titles

    # ---------- Modules → Discussion mapping
    def get_modules(self, course_id: str) -> List[dict]:
        url = f'{self.get_server_url()}api/v1/courses/{course_id}/modules?per_page=100'
        return self._paged_get(url)

    def get_module_items(self, course_id: str, module_id: int) -> List[dict]:
        url = f'{self.get_server_url()}api/v1/courses/{course_id}/modules/{module_id}/items?per_page=100'
        return self._paged_get(url)

    def build_module_discussion_map(self, course_id: str) -> None:
        self.module_discussion_map = {}
        modules = self.get_modules(course_id)
        # Debug: surface empty modules early
        if not modules:
            print("No Canvas modules returned (empty list). Check course has Modules enabled and API token scope.")
            return
        for mod in modules:
            module_name = (mod.get('name') or f"Module {mod.get('id', 'Unknown')}").strip()
            module_id = mod.get('id')
            if not isinstance(module_id, int):
                continue
            items = self.get_module_items(course_id, module_id)
            # Optional debug:
            # print(f"Module '{module_name}': {len(items)} items")
            for it in items:
                if it.get('type') == 'Discussion':
                    self.module_discussion_map.setdefault(module_name,
                                                          []).append(it['content_id'])

    # ---------- XLSX Writer
    def _output_basepath(self) -> Path:
        download_folder = Path.home() / 'Downloads'
        download_folder.mkdir(parents=True, exist_ok=True)
        return download_folder

    def _role_label(self) -> str:
        return "students" if self.get_enrollment_type() == "StudentEnrollment" else "instructors"

    def _unique_sheet_name(self, wb: Workbook, base: str) -> str:
        invalid = set('[]:*?/\\')
        cleaned = ''.join(ch for ch in base if ch not in invalid).strip().rstrip('.')
        if not cleaned:
            cleaned = "Sheet"
        name = cleaned[:31]
        if name not in wb.sheetnames:
            return name
        n = 2
        while True:
            candidate = f"{name[:31 - len(str(n)) - 1]} {n}"
            if candidate not in wb.sheetnames:
                return candidate
            n += 1

    def _autowidth(self, ws, max_rows: int, max_cols: int) -> None:
        for c in range(1, max_cols + 1):
            col_letter = get_column_letter(c)
            max_len = 10
            for r in range(1, max_rows + 1):
                val = ws.cell(row=r, column=c).value
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    def write_module_breakdown_xlsx(self, enrollee_discussion_data: Dict[str, List[bool]]) -> None:
        # Why: Users sometimes call this directly; ensure module map exists.
        if not self.module_discussion_map:
            # Attempt to build with the current course id
            self.build_module_discussion_map(self.course_num)

        if not self.discussions_meta:
            print("No discussion data to write.")
            return

        id_to_index: Dict[int, int] = {int(meta["id"]): i for i, meta in enumerate(self.discussions_meta) if isinstance(meta.get("id"), int)}

        wb = Workbook()
        wb.remove(wb.active)

        # ---- Overview_All (all discussions)
        overview = wb.create_sheet(title="Overview_All")
        all_headers = ["Name"] + [str(meta["title"]) for meta in self.discussions_meta]
        for j, header in enumerate(all_headers, start=1):
            overview.cell(row=1, column=j, value=header)
        r = 2
        for enrollee, participations in enrollee_discussion_data.items():
            overview.cell(row=r, column=1, value=enrollee)
            for j, p in enumerate(participations, start=2):
                overview.cell(row=r, column=j, value="Yes" if p else "No")
            r += 1
        self._autowidth(overview, r - 1, len(all_headers))

        # ---- Per-module sheets (Discussion items only)
        if not self.module_discussion_map:
            print("No per-module sheets written (module map is empty).")
        for module_name in sorted(self.module_discussion_map.keys()):
            topic_ids = self.module_discussion_map.get(module_name, [])
            col_indices: List[int] = [id_to_index[tid] for tid in topic_ids if tid in id_to_index]
            if not col_indices:
                continue
            ws = wb.create_sheet(title=self._unique_sheet_name(wb, module_name))
            headers = ["Name"] + [str(self.discussions_meta[i]["title"]) for i in col_indices]
            for j, header in enumerate(headers, start=1):
                ws.cell(row=1, column=j, value=header)
            r = 2
            for enrollee, participations in enrollee_discussion_data.items():
                ws.cell(row=r, column=1, value=enrollee)
                for k, overall_idx in enumerate(col_indices, start=2):
                    ws.cell(row=r, column=k, value="Yes" if participations[overall_idx] else "No")
                r += 1
            self._autowidth(ws, r - 1, len(headers))

        role = self._role_label()
        xlsx_path = self._output_basepath() / f"{self.course_name}__{role}_discussions.xlsx"
        wb.save(xlsx_path)
        print(f"XLSX written: {xlsx_path}")

    # ---------- Course name
    def set_course_name(self, course_id: str) -> None:
        course_url = f'{self.get_server_url()}api/v1/courses/{course_id}'
        response = requests.get(course_url, headers=self.headers())
        try:
            course = response.json()
            self.course_name = course.get('name', 'Unknown Course')
        except Exception:
            self.course_name = 'Unknown Course'

    def get_course_name(self) -> str:
        return self.course_name


def main() -> None:
    course_num = "1645103"
    canvas = CanvasDiscussions('LPS_Production', 'Student', course_num)
    canvas.set_course_name(course_num)
    print(f"Course Name: {canvas.course_name}")

    course_enrollees = canvas.get_enrollees(course_num)
    if canvas.get_enrollment_type() != 'StudentEnrollment':
        if canvas.get_enrollment_type() == 'TaEnrollment':
            canvas.set_enrollment_type('Teacher')
        else:
            canvas.set_enrollment_type('TA')
        course_enrollees += canvas.get_enrollees(course_num)

    if course_enrollees:
        data, _titles = canvas.get_course_discussion_data(course_num, course_enrollees)
        if data:
            canvas.write_module_breakdown_xlsx(data)
            return
    print(f"No XLSX written for {canvas.course_name}")


if __name__ == '__main__':
    main()
